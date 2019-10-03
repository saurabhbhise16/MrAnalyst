import tkinter as tk
from tkinter import Toplevel, Button, Menu
from tkinter import font  as tkfont
from fileinput import filename
from tkinter.filedialog import askopenfilename
import pandas as pd
import xlrd
import csv
import matplotlib.pyplot as plt
import os
#import openpyxl
#from tkinter import *
#from openpyxl import load_workbook
from openpyxl import *
from tkinter import *


#m = tk.Tk()

#method to drag file
#def chooseFile():
#    from tkinter.filedialog import askopenfilename
#    filename = askopenfilename()
#    print(filename)

#method for converting excel file to csv
#def csv_from_excel():
#    wb = xlrd.open_workbook('newdata.xlsx')
#    sh = wb.sheet_by_name('Sheet1')
#    new_csv_file = open('your_csv_file.csv', 'w')
#    wr = csv.writer(new_csv_file, quoting=csv.QUOTE_ALL)
#    for rownum in range(sh.nrows):
#        wr.writerow(sh.row_values(rownum))
#    new_csv_file.close()


#method for printing entered value
#def hi():
#    print(startEntry.get())

#Actual Program Starts Here

#chooseFile()

#startLabel = tk.Label(m, text = "Enter Text: ")
#startLabel.pack()

#startEntry = tk.Entry(m)
#startEntry.pack()

#plotButton = tk.Button(m,text="Enter", command=hi)
#plotButton.pack()



#csv_from_excel()
#m.mainloop()


class Application(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.title_font = tkfont.Font(family='Helvetica', size=18, weight="bold", slant="italic")
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (MainMenu, MyStudents, Marks):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("MainMenu")

    def show_frame(self, page_name):
        #Show a frame for the given page name
        frame = self.frames[page_name]
        frame.tkraise()


class MainMenu(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Mr. Analyst", font=controller.title_font)
        label.pack(side="top", fill="x", pady=10)

        button1 = tk.Button(self, text="My Students", command=lambda: controller.show_frame("MyStudents"))
        button2 = tk.Button(self, text="Marks Evaluation", command=lambda: controller.show_frame("Marks"))
        button1.pack()
        button2.pack()


class MyStudents(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        wb = load_workbook('/home/honey/PycharmProjects/miniproject/test.xlsx')
        self.sheet = wb.active
        self.controller = controller
        label = tk.Label(self, text="Student Profiles", font=controller.title_font)
        label.pack(side="top", fill="x", pady=10)
        self.newstudent = tk.Button(self, text="New Registrations", command=new_registration)
        self.newstudent.pack()
        #self.profile = tk.Button(self, text="Student Profiles", command=lambda: controller.show_frame(""))
        #self.profile.pack()
        button = tk.Button(self, text="Main Menu", command=lambda: controller.show_frame("MainMenu"))
        button.pack()

class new_registration:
    def __init__(self):
        self.wb = load_workbook('/home/honey/PycharmProjects/miniproject/test.xlsx')
        self.sheet = self.wb.active
        root = Tk()
        root.configure(background='light green')
        root.title("registration form")
        root.geometry("500x300")
        self.excel()
        heading = Label(root, text="Form", bg="light green")
        name = Label(root, text="Name", bg="light green")
        course = Label(root, text="Course", bg="light green")
        sem = Label(root, text="Semester", bg="light green")
        form_no = Label(root, text="Form No.", bg="light green")
        contact_no = Label(root, text="Contact No.", bg="light green")
        email_id = Label(root, text="Email id", bg="light green")
        address = Label(root, text="Address", bg="light green")

        heading.grid(row=0, column=1)
        name.grid(row=1, column=0)
        course.grid(row=2, column=0)
        sem.grid(row=3, column=0)
        form_no.grid(row=4, column=0)
        contact_no.grid(row=5, column=0)
        email_id.grid(row=6, column=0)
        address.grid(row=7, column=0)

        self.name_field = Entry(root)
        self.course_field = Entry(root)
        self.sem_field = Entry(root)
        self.form_no_field = Entry(root)
        self.contact_no_field = Entry(root)
        self.email_id_field = Entry(root)
        self.address_field = Entry(root)
        self.name_field.grid(row=1, column=1, ipadx="100")
        self.course_field.grid(row=2, column=1, ipadx="100")
        self.sem_field.grid(row=3, column=1, ipadx="100")
        self.form_no_field.grid(row=4, column=1, ipadx="100")
        self.contact_no_field.grid(row=5, column=1, ipadx="100")
        self.email_id_field.grid(row=6, column=1, ipadx="100")
        self.address_field.grid(row=7, column=1, ipadx="100")

        # call excel function
        self.excel()

        submit = Button(root, text="Submit", fg="Black", bg="Red", command=self.insert)
        submit.grid(row=8, column=1)

        # start the GUI
        root.mainloop()

    def excel(self):
        # resize the width of columns in
        # excel spreadsheet
        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 10
        self.sheet.column_dimensions['C'].width = 10
        self.sheet.column_dimensions['D'].width = 20
        self.sheet.column_dimensions['E'].width = 20
        self.sheet.column_dimensions['F'].width = 40
        self.sheet.column_dimensions['G'].width = 50

        self.sheet.cell(row=1, column=1).value = "Name"
        self.sheet.cell(row=1, column=2).value = "Course"
        self.sheet.cell(row=1, column=3).value = "Semester"
        self.sheet.cell(row=1, column=4).value = "Form Number"
        self.sheet.cell(row=1, column=5).value = "Contact Number"
        self.sheet.cell(row=1, column=6).value = "Email id"
        self.sheet.cell(row=1, column=7).value = "Address"

    def clear(self):
        # clear the content of text entry box
        self.name_field.delete(0, END)
        self.course_field.delete(0, END)
        self.sem_field.delete(0, END)
        self.form_no_field.delete(0, END)
        self.contact_no_field.delete(0, END)
        self.email_id_field.delete(0, END)
        self.address_field.delete(0, END)

    def insert(self):
        if (self.name_field.get() == "" and
                self.course_field.get() == "" and
                self.sem_field.get() == "" and
                self.form_no_field.get() == "" and
                self.contact_no_field.get() == "" and
                self.email_id_field.get() == "" and
                self.address_field.get() == ""):

            print("empty input")

        else:
            current_row = self.sheet.max_row
            self.sheet.cell(row=current_row + 1, column=1).value = self.name_field.get()
            self.sheet.cell(row=current_row + 1, column=2).value = self.course_field.get()
            self.sheet.cell(row=current_row + 1, column=3).value = self.sem_field.get()
            self.sheet.cell(row=current_row + 1, column=4).value = self.form_no_field.get()
            self.sheet.cell(row=current_row + 1, column=5).value = self.contact_no_field.get()
            self.sheet.cell(row=current_row + 1, column=6).value = self.email_id_field.get()
            self.sheet.cell(row=current_row + 1, column=7).value = self.address_field.get()

            # save the file
            self.wb.save('/home/honey/PycharmProjects/miniproject/test.xlsx')

            # set focus on the name_field box

            # call the clear() function
            self.clear()


class Marks(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Marks Evaluation", font=controller.title_font)
        label.pack(side="top", fill="x", pady=10)
        self.UT_marks = tk.Button(self, text="UT-1 Marks", command= self.marksUT)
        self.UT_marks.pack()
        button = tk.Button(self, text="Main Menu",command=lambda: controller.show_frame("MainMenu"))
        button.pack()

    def marksUT(self):
        df = pd.read_csv("fakedata.csv", sep=",").set_index("Roll_no")
        d = dict(zip(df.index, df.values.tolist()))
        print(d)
        df.set_index('student')[['sub1', 'sub2', 'sub3']].plot.bar()
        plt.show()

if __name__ == "__main__":
    app = Application()
    app.mainloop()




