'''# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

class abc:
    def __init__(self):
        self.wb = load_workbook('/home/honey/PycharmProjects/miniproject/test.xlsx')
        self.sheet = self.wb.active
        root = Tk()
        root.configure(background='light green')
        root.title("registration form")
        root.geometry("500x300")
        self.excel()
        heading = Label(root, text="Form", bg="light green")
        name = Label(root, text="Name", bg="light green")
        course = Label(root, text="Course", bg="light green")
        sem = Label(root, text="Semester", bg="light green")
        form_no = Label(root, text="Form No.", bg="light green")
        contact_no = Label(root, text="Contact No.", bg="light green")
        email_id = Label(root, text="Email id", bg="light green")
        address = Label(root, text="Address", bg="light green")

        heading.grid(row=0, column=1)
        name.grid(row=1, column=0)
        course.grid(row=2, column=0)
        sem.grid(row=3, column=0)
        form_no.grid(row=4, column=0)
        contact_no.grid(row=5, column=0)
        email_id.grid(row=6, column=0)
        address.grid(row=7, column=0)

        self.name_field = Entry(root)
        self.course_field = Entry(root)
        self.sem_field = Entry(root)
        self.form_no_field = Entry(root)
        self.contact_no_field = Entry(root)
        self.email_id_field = Entry(root)
        self.address_field = Entry(root)
        self.name_field.grid(row=1, column=1, ipadx="100")
        self.course_field.grid(row=2, column=1, ipadx="100")
        self.sem_field.grid(row=3, column=1, ipadx="100")
        self.form_no_field.grid(row=4, column=1, ipadx="100")
        self.contact_no_field.grid(row=5, column=1, ipadx="100")
        self.email_id_field.grid(row=6, column=1, ipadx="100")
        self.address_field.grid(row=7, column=1, ipadx="100")

        # call excel function
        self.excel()

        submit = Button(root, text="Submit", fg="Black", bg="Red", command=self.insert)
        submit.grid(row=8, column=1)

        # start the GUI
        root.mainloop()

    def excel(self):
        # resize the width of columns in
        # excel spreadsheet
        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 10
        self.sheet.column_dimensions['C'].width = 10
        self.sheet.column_dimensions['D'].width = 20
        self.sheet.column_dimensions['E'].width = 20
        self.sheet.column_dimensions['F'].width = 40
        self.sheet.column_dimensions['G'].width = 50

        self.sheet.cell(row=1, column=1).value = "Name"
        self.sheet.cell(row=1, column=2).value = "Course"
        self.sheet.cell(row=1, column=3).value = "Semester"
        self.sheet.cell(row=1, column=4).value = "Form Number"
        self.sheet.cell(row=1, column=5).value = "Contact Number"
        self.sheet.cell(row=1, column=6).value = "Email id"
        self.sheet.cell(row=1, column=7).value = "Address"

    def clear(self):
        # clear the content of text entry box
        self.name_field.delete(0, END)
        self.course_field.delete(0, END)
        self.sem_field.delete(0, END)
        self.form_no_field.delete(0, END)
        self.contact_no_field.delete(0, END)
        self.email_id_field.delete(0, END)
        self.address_field.delete(0, END)

    def insert(self):
        if (self.name_field.get() == "" and
                self.course_field.get() == "" and
                self.sem_field.get() == "" and
                self.form_no_field.get() == "" and
                self.contact_no_field.get() == "" and
                self.email_id_field.get() == "" and
                self.address_field.get() == ""):

            print("empty input")

        else:
            current_row = self.sheet.max_row
            self.sheet.cell(row=current_row + 1, column=1).value = self.name_field.get()
            self.sheet.cell(row=current_row + 1, column=2).value = self.course_field.get()
            self.sheet.cell(row=current_row + 1, column=3).value = self.sem_field.get()
            self.sheet.cell(row=current_row + 1, column=4).value = self.form_no_field.get()
            self.sheet.cell(row=current_row + 1, column=5).value = self.contact_no_field.get()
            self.sheet.cell(row=current_row + 1, column=6).value = self.email_id_field.get()
            self.sheet.cell(row=current_row + 1, column=7).value = self.address_field.get()

            # save the file
            self.wb.save('/home/honey/PycharmProjects/miniproject/test.xlsx')

            # set focus on the name_field box

            # call the clear() function
            self.clear()


if __name__ == "__main__":
    # create a GUI window
    abc()'''

import image
import tkinter


window = tkinter.Tk()

window.title("Join")
window.geometry("300x300")
window.configure(background='grey')

imageFile = "image.jpg"

window.im1 = image.open(imageFile)


#raw_input()
window.mainloop()